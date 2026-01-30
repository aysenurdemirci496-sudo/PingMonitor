import json
import os
from openpyxl import load_workbook

DEVICES_JSON = "devices.json"
DEVICES_XLSX = "devices.xlsx"



def add_device_to_excel(device):
    from openpyxl import load_workbook

    wb = load_workbook("devices.xlsx")
    ws = wb.active

    ws.append([
        device.get("name"),
        device.get("ip"),
        device.get("device"),
        device.get("model"),
        device.get("mac"),
        device.get("location"),
        device.get("unit"),
        device.get("description")
    ])

    wb.save("devices.xlsx")


def load_devices():
    if os.path.exists(DEVICES_JSON):
        try:
            with open(DEVICES_JSON, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def load_devices_from_excel(path, mapping):
    devices = []

    # 1️⃣ ÖNCE PANDAS DENE
    try:
        import pandas as pd

        df = pd.read_excel(path)

        for _, row in df.iterrows():
            device = {}
            for field, header in mapping.items():
                device[field] = row.get(header)
            devices.append(device)

        return devices

    except Exception:
        pass   # pandas yoksa sessizce devam et

    # 2️⃣ FALLBACK → OPENPYXL
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    header_index = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        device = {}

        for field, header in mapping.items():
            idx = header_index.get(header)
            device[field] = row[idx] if idx is not None else None

        devices.append(device)

    return devices
def save_devices(devices):
    with open(DEVICES_JSON, "w", encoding="utf-8") as f:
        json.dump(devices, f, indent=2, ensure_ascii=False)


def update_device_in_excel(old_ip, updated_device, xlsx_file=DEVICES_XLSX):
    if not os.path.exists(xlsx_file):
        return

    wb = load_workbook(xlsx_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(old_ip):
            row[0].value = updated_device["name"]
            row[1].value = updated_device["ip"]
            row[2].value = updated_device.get("device", "")
            row[3].value = updated_device.get("model", "")
            row[4].value = updated_device.get("mac", "")
            row[5].value = updated_device.get("location", "")
            row[6].value = updated_device.get("unit", "")
            row[7].value = updated_device.get("description", "")
            break

    wb.save(xlsx_file)